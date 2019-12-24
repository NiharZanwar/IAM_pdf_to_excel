import pdftotext
import os
import openpyxl


work_dir = os.getcwd()
list_dir = os.listdir(work_dir)

def init():
    if 'pdf' not in list_dir:
        os.mkdir(work_dir + '/pdf')
        os.mkdir(work_dir + '/pdf/' + 'uploaded_pdf_ack')
        os.mkdir(work_dir + '/pdf/' + 'uploaded_pdf_invoice')


date_dict = {"January": "01",
             "February": "02",
             "March": "03",
             "April": "04",
             "May": "05",
             "June": "06",
             "July": "07",
             "August": "08",
             "September": "09",
             "October": "10",
             "November": "11",
             "December": "12"}


pdf_dir = work_dir + '/pdf'


def make_log(log):
    log_str = str(log)
    with open('log.txt', 'a') as f:
        f.write(log_str)


def remove_file(filename):
    try:
        os.remove(pdf_dir + '/' + filename)
        return
    except:
        return


def convert_str_date(string):
    string = string.replace(',', '')
    date_list = string.split(' ')
    month = date_dict[date_list[0]]
    rcv_date = date_list[1] + '-' + month + '-' + date_list[2]
    return rcv_date


def get_bill_details(pdf):
    lines = pdf.split('\n')
    bill_no = ''
    invoice_date = ''
    gross_amount = ''
    less_tds = ''
    less_tds_work = ''
    net_amount = ''

    rcv_date = convert_str_date(lines[1])
    make_log(rcv_date)

    for i in range(0, len(lines)):
        if 'Details of Bill' in lines[i]:
            details = lines[i+1]
            details = details.replace(' ', '').split('/')
            bill_no = details[0]
            make_log(bill_no)
            invoice_date = details[1].replace('.', '-')
            make_log(invoice_date)
            continue

        if 'Gross Amount' in lines[i]:
            gross_amount = float(((((lines[i].replace(' ', '')).split(':'))[1]).replace('*', '')).replace(',', ''))
            make_log(gross_amount)
            less_tds = float((lines[i+1].replace(' ', '').split(':'))[1].replace(',', ''))
            make_log(less_tds)
            less_tds_work = float((lines[i+2].replace(' ', '').split(':'))[1].replace(',', ''))
            make_log(less_tds_work)
            net_amount = float(((((lines[i+3].replace(' ', '')).split(':'))[1]).replace('*', '')).replace(',', ''))
            make_log(net_amount)
    return net_amount, gross_amount, invoice_date, less_tds, less_tds_work, rcv_date, bill_no


def add_to_excel(filename):
    with open(work_dir + '/pdf/' + filename, "rb") as f:
        make_log(filename + '\n')
        pdf = pdftotext.PDF(f)
        net_amount, gross_amount, invoice_date, less_tds, less_tds_work, rcv_date, bill_no = get_bill_details(pdf[0])
        br = '<br>'
        hr = '<hr>'
        to_return = filename + br + 'Receive Date : ' + rcv_date + br + 'Invoice Date : ' + invoice_date + br + \
                    'Invoice Number : ' + bill_no + br  + 'Gross Amount : ' + str(gross_amount) + br + 'Less TDS : ' + \
                    str(less_tds) + br + 'Less TSD WORK : ' + str(less_tds_work) + br + 'Net Amount : ' + str(net_amount)\
                    + hr

        book = openpyxl.load_workbook('invoice.xlsx')
        sheet = book.active

        invoice_repeat = False

        for i in range(2, sheet.max_row + 1):
            if sheet['B' + str(i)].value == bill_no:
                make_log('REPEAT DETECTED')
                invoice_repeat = True

        if invoice_repeat:
            remove_file(filename)
            book.close()
            return 'Repeated : ' + filename + hr

        else:
            maximum_row = sheet.max_row
            sheet['A' + str(maximum_row + 1)] = invoice_date
            sheet['B' + str(maximum_row + 1)] = bill_no
            sheet['C' + str(maximum_row + 1)] = gross_amount
            sheet['D' + str(maximum_row + 1)] = less_tds
            sheet['E' + str(maximum_row + 1)] = less_tds_work
            sheet['F' + str(maximum_row + 1)] = net_amount
            sheet['G' + str(maximum_row + 1)] = rcv_date

        book.save('invoice.xlsx')
        book.close()

        os.rename(pdf_dir + '/' + filename, pdf_dir + '/uploaded_pdf_invoice/' + bill_no + '.pdf')
        return to_return


def get_ack_details(pdf):
    lines = pdf.split('\n')
    inv_number = ''
    inv_date = ''
    inv_amount = 0.0
    po_number = ''
    ref_number = ''
    for i in range(0, len(lines)):

        if 'Invoice Number' in lines[i]:
            inv_number = (lines[i].replace(' ', '')).split(':')[1]
            make_log(inv_number)
            continue
        if 'Invoice Date' in lines[i]:
            inv_date = (lines[i].replace(' ', '').split(':'))[1].replace('.', '-')
            make_log(inv_date)
            continue
        if 'Invoice Amount' in lines[i]:
            inv_amount = float((lines[i].replace(' ', '').split(':')[1]).replace(',', ''))
            make_log(str(inv_amount))
            continue
        if 'PO Number' in lines[i]:
            po_number = (lines[i].replace(' ', '').split(':')[1])
            make_log(po_number)
            continue
        if 'Reference Number' in lines[i]:
            ref_number = lines[i].replace(' ', '').split(':')[1]
            make_log(ref_number)

    return inv_number, inv_date, inv_amount, po_number, ref_number


def add_ack_toexcel(filename):

    with open(work_dir + '/pdf/' + filename, "rb") as f:
        make_log(filename + '\n')
        pdf = pdftotext.PDF(f)
        inv_number, inv_date, inv_amount, po_number, ref_number = get_ack_details(pdf[0])

        br = '<br>'
        hr = '<hr>'
        to_return = filename + br + 'Invoice Number : ' + inv_number + br + 'Invoice Date : ' + inv_date + br + \
                    'Invoice Amount : ' + str(inv_amount) + br + 'PO number : ' + po_number + br + 'Reference Number : ' + \
                    ref_number + hr

        book = openpyxl.load_workbook('acknowledgement.xlsx')
        sheet = book.active

        invoice_repeat = False

        for i in range(2, sheet.max_row + 1):
            if sheet['B' + str(i)].value == inv_number:
                make_log('REPEAT DETECTED')
                invoice_repeat = True

        if invoice_repeat:
            remove_file(filename)
            book.close()
            return 'Repeated : ' + filename + hr
        else:
            maximum_row = sheet.max_row
            sheet['A' + str(maximum_row + 1)] = inv_date
            sheet['B' + str(maximum_row + 1)] = inv_number
            sheet['C' + str(maximum_row + 1)] = inv_amount
            sheet['D' + str(maximum_row + 1)] = po_number
            sheet['E' + str(maximum_row + 1)] = ref_number

        book.save('acknowledgement.xlsx')
        book.close()

        os.rename(pdf_dir + '/' + filename, pdf_dir + '/uploaded_pdf_ack/' + inv_number + '.pdf')
        make_log('----------------------------------')
        return to_return


